import sys
import os
import re



import openpyxl
from langchain_google_genai import ChatGoogleGenerativeAI
from langgraph.prebuilt import create_react_agent
from langgraph.checkpoint.memory import MemorySaver

def calculate_irr(cash_flows, iterations=1000):
    """Calculates the Internal Rate of Return (IRR) using Newton's method."""
    rate = 0.1  # Starting guess
    for _ in range(iterations):
        npv = sum(cf / (1 + rate)**i for i, cf in enumerate(cash_flows))
        derivative = sum(-i * cf / (1 + rate)**(i + 1) for i, cf in enumerate(cash_flows))
        if derivative == 0:
            break
        new_rate = rate - npv / derivative
        if abs(new_rate - rate) < 1e-7:
            return new_rate
        rate = new_rate
    return rate

def get_revenue_estimate() -> str:
    """Reads the Q4 revenue estimate from q4_data.txt.
    
    Returns:
        The content of the q4_data.txt file.
    """
    adk_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'adk'))
    q4_data_path = os.path.join(adk_dir, 'q4_data.txt')
    with open(q4_data_path, 'r') as f:
        return f.read()

def update_financial_model(revenue: float) -> str:
    """Updates the Year 5 revenue in mock_model.xlsx and recalculates IRR.
    
    Args:
        revenue: The new revenue value for Year 5.
        
    Returns:
        A message containing the updated IRR.
    """
    adk_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'adk'))
    mock_model_path = os.path.join(adk_dir, 'mock_model.xlsx')
    wb = openpyxl.load_workbook(mock_model_path, data_only=False)
    ws = wb['DCF Model']
    
    # Update Year 5 Revenue (Cell F2)
    ws['F2'] = revenue
    
    # Calculate IRR dynamically.
    initial_investment = ws['B5'].value
    
    cash_flows = [initial_investment]
    for col in range(2, 7):  # B to F
        col_revenue = ws.cell(row=2, column=col).value
        net_cash_flow = float(col_revenue) * 0.7
        cash_flows.append(net_cash_flow)
    
    irr = calculate_irr(cash_flows)
    
    adk_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'adk'))
    updated_model_path = os.path.join(adk_dir, 'mock_model_updated.xlsx')
    wb.save(updated_model_path)
    
    return f"Excel model updated with revenue ${revenue:,.2f}. The calculated IRR is approximately {irr:.2%}."

def write_summary_report(summary: str) -> str:
    """Writes a short summary report to summary.md.
    
    Args:
        summary: The text content for the summary report.
        
    Returns:
        A success message.
    """
    adk_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'adk'))
    summary_path = os.path.join(adk_dir, 'summary.md')
    with open(summary_path, 'w') as f:
        f.write(summary)
    return "Summary report written to summary.md."

# Initialize the model using Vertex AI
llm = ChatGoogleGenerativeAI(
    model="gemini-2.5-flash",
    vertexai=True,
    project=os.environ.get('GOOGLE_CLOUD_PROJECT'),
    location=os.environ.get('GOOGLE_CLOUD_LOCATION', 'us-central1')
)

# Define tools
tools = [get_revenue_estimate, update_financial_model, write_summary_report]

# Create the LangGraph ReAct agent
system_message = """
You are a financial analyst agent. Your task is to:
1. Read the Q4 revenue estimate from 'q4_data.txt' using get_revenue_estimate.
2. Extract the numeric revenue value from the text.
3. Update the existing Excel spreadsheet 'mock_model.xlsx' with this new revenue using update_financial_model.
4. Based on the new IRR reported by the tool, draft a short summary of the update in 'summary.md' using write_summary_report.

Be concise in your summary.
"""

financial_langgraph_agent = create_react_agent(llm, tools, checkpointer=MemorySaver(), prompt=system_message)
