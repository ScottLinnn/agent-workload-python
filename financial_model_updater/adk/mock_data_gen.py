import sys
import os

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def generate_mock_data():
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    q4_data_path = os.path.join(base_path, 'q4_data.txt')
    excel_path = os.path.join(base_path, 'mock_model.xlsx')

    # 1. Create q4_data.txt
    revenue_estimate = 1500000
    with open(q4_data_path, 'w') as f:
        f.write(f"Q4 Revenue Estimate: ${revenue_estimate:,}\n")
    
    # 2. Create mock_model.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    
    # Header
    ws['A1'] = "Year"
    ws['B1'] = "Year 1"
    ws['C1'] = "Year 2"
    ws['D1'] = "Year 3"
    ws['E1'] = "Year 4"
    ws['F1'] = "Year 5"
    
    # Revenue (B2:F2)
    ws['A2'] = "Revenue"
    ws['B2'] = 1000000
    ws['C2'] = 1100000
    ws['D2'] = 1210000
    ws['E2'] = 1331000
    ws['F2'] = 1464100 # This will be updated by the agent
    
    # Expenses (30% of Revenue)
    ws['A3'] = "Expenses"
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}3'] = f"={col_letter}2 * 0.3"
    
    # Net Cash Flow
    ws['A4'] = "Net Cash Flow"
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}4'] = f"={col_letter}2 - {col_letter}3"
    
    # Initial Investment
    ws['A5'] = "Initial Investment"
    ws['B5'] = -3000000
    
    # IRR calculation (using a range B4:F4 + B5)
    # Note: simple IRR in Excel formula
    ws['A7'] = "IRR"
    ws['B7'] = "=IRR(B5:F4)" # Range includes initial investment and subsequent cash flows
    
    wb.save(excel_path)
    print(f"Generated {q4_data_path} and {excel_path}")

if __name__ == "__main__":
    generate_mock_data()
